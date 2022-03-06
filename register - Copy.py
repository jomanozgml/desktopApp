# -*- coding: utf-8 -*-
import os
#Importing all from Tkinter for GUI
from Tkinter import *
#Importing openFileDialog and saveFileDialog
from tkFileDialog import *
from tkMessageBox import *
import PIL
from PIL import Image, ImageTk
import numpy as np
import cv2
import re
# from openpyxl import Workbook
from openpyxl import load_workbook

window, dashWindow, searchEntry = None, None, None

loginUser, loginPass = None, None

userGroup = ['nanc', 'manoj', 'shova']
passGroup = ['register', 'shrestha', 'mangal']
userDictionary = {'nanc':'register', 'manoj':'shrestha', 'shova':'mangal'}


def dataEntry(prevWindow, opt):
	newEntryBtn, wb, ws = None, None, None
	e1, e2, e3, e4, e5, e6, e7, e12, e13 = [], None, [], None, None, None, None, None, None
	c8, c9, c10, c11, tkvar, tkvar2, tkvar3 = None, None, None, None, [], None, None
	i, a, lastRow = 0, 0, 0
	snValue, ckvar8, ckvar9, ckvar10, ckvar11, editSaveBtn = [], [], [], [], [], []
	titleName = ['S.N.', 'Date', 'Category', 'Name', 'Expenses Type', 'Site', 'Duration',\
		'Status', '', '', '', 'Issue', 'Remarks']
	entryName = ['']
	rowsDash = 1
	colsDash = len(titleName)
	
	#Loading Excel WorkBook
	global wb, ws, lastRow
	wb = load_workbook('reg.xlsx')
	ws = wb['Sheet']
	lastRow = ws.max_row
	
	if opt == 1:
		print i
		newWindow = Frame(mainWindow, bg="#cfd8dc")
		newWindow.place(anchor=NW, width=1200, height=800)
		
	if opt == 2:
		print i
		global snValue
		
		searchResult = None
		searchThis = searchEntry.get()
		searchRegex = re.compile(r'.*' + re.escape(searchThis) + r'.*', re.I)
		for rowNum in range(2, lastRow+1):
			if searchThis != None:
				searchName = ws.cell(row=rowNum, column=4).value
				searchResult = searchRegex.search(searchName)
				
			if  searchThis == None or searchResult != None:
					snValue.append(ws.cell(row=rowNum, column=1).value)			
		
		def onFrameConfigure(canvas):
			#Reset the scroll region to encompass the inner frame
			canvas.configure(scrollregion=canvas.bbox("all"))
	
		canvas = Canvas(prevWindow, width=1200, height=600)
		newWindow = Frame(canvas, bg="#cfd8dc")
		scrollBar = Scrollbar(prevWindow, orient="vertical", command=canvas.yview)
		canvas.configure(yscrollcommand=scrollBar.set)
		scrollBar.pack(side="right", fill="y")
		canvas.pack(side="left", fill="both", expand=True)
		canvas.create_window((4,4), window=newWindow, anchor="nw")
		newWindow.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))
	
	for j in range(0, colsDash): #Columns
		b = Label(newWindow, text=titleName[j], padx=5, pady=5, bd=1, relief="flat", bg="#cfd8dc")
		b.grid(row=0, column=j)
		ws.cell(row=1, column=j+1, value=titleName[j])
	
	def newEntry():
		global i, a, newEntryBtn, e1, e2, e3, tkvar, e4, e5, tkvar2, e6, e7, ckvar8,\
		ckvar9, ckvar10, ckvar11, e12, tkvar3, e13, c8, c9, c10, c11, editSaveBtn, lastRow
		i = i+1
			
		e1.append(Entry(newWindow, relief="flat", width=4, bg="#f0f0f0", bd=5, justify=CENTER))
		e1[a].grid(row=i, column=0)
		
		if opt == 1:
			e1[a].insert(0,lastRow)
				
		e2 = Entry(newWindow, relief="flat", width=9, bg="#f0f0f0", bd=5, justify=CENTER)
		e2.grid(row=i, column=1)
				
		tkvar.append(StringVar())
		categoryDash = {'Employee','Contractor','Supplier','Customer'}
		tkvar[a].set('Employee')
		e3.append(OptionMenu(newWindow, tkvar[a], *categoryDash))
		e3[a].config(width=8)
		e3[a].grid(row=i, column=2)
		
		e4 = Entry(newWindow, relief="flat", width=20, bg="#f0f0f0", bd=5, justify=CENTER)
		e4.grid(row=i, column=3)
				
		tkvar2 = StringVar()
		expensesDash = {'Site Expenses','Personal Expenses','Labour Expenses','Mess Expenses', 'Advance'}
		tkvar2.set('Site Expenses')
		e5 = OptionMenu(newWindow, tkvar2, *expensesDash)
		e5.config(width=14)
		e5.grid(row=i, column=4)
				
		e6 = Entry(newWindow, relief="flat", width=8, bg="#f0f0f0", bd=5, justify=CENTER)
		e6.grid(row=i, column=5)
				
		e7 = Entry(newWindow, relief="flat", width=9, bg="#f0f0f0", bd=5, justify=CENTER)
		e7.grid(row=i, column=6)
				
		ckvar8.append(StringVar())
		c8 = Checkbutton(newWindow, text="Recorded", onvalue="Recorded", offvalue=None, variable=ckvar8[a])
		c8.grid(row=i, column=7)
		c8.deselect()
		
		ckvar9.append(StringVar())
		c9 = Checkbutton(newWindow, text="Checked", onvalue="Checked", offvalue=None, variable=ckvar9[a])
		c9.grid(row=i, column=8)
		c9.deselect()
		
		ckvar10.append(StringVar())
		c10 = Checkbutton(newWindow, text="Pending", onvalue="Pending", variable=ckvar10[a])
		c10.grid(row=i, column=9)
		c10.deselect()
		
		ckvar11.append(StringVar())
		c11 = Checkbutton(newWindow, text="Completed", onvalue="Completed", variable=ckvar11[a])
		c11.grid(row=i, column=10)
		c11.deselect()
		
		tkvar3 = StringVar()
		issues = {'ROM Missing','Bill Missing',"Doc Didn't Match",'Shipping Missing', 'Other'}
		tkvar3.set('')
		e12 = OptionMenu(newWindow, tkvar3, *issues)
		e12.config(width=14)
		e12.grid(row=i, column=11)
		
		e13 = Entry(newWindow, relief="flat", width=24, bg="#f0f0f0", bd=5, justify=CENTER)
		e13.grid(row=i, column=12)
		
		if opt == 1:
			newEntryBtn = Button(newWindow, text=u"\u271a", command=lambda: saveData(1))
			newEntryBtn.grid(row=i, column=13)
		if opt == 2:
			editSaveBtn.append(Button(newWindow, text=u"\u270e"))
			editSaveBtn[a].grid(row=i, column=13)
		#text=u"\u2714"  text=u"\u2718"
		
		a = a+1
		lastRow = lastRow+1
		
	if opt == 1:
		newEntry()
	
	if opt == 2:
		if snValue:
			snValueLen = len(snValue)
			for x in range(0, snValueLen):
				newEntry()
				e1[a-1].insert(0, snValue[x])
				e2.insert(0, ws.cell(row=snValue[x]+1, column=2).value)
				tkvar[a-1].set(ws.cell(row=snValue[x]+1, column=3).value)
				e4.insert(0, ws.cell(row=snValue[x]+1, column=4).value)
				tkvar2.set(ws.cell(row=snValue[x]+1, column=5).value)
				e6.insert(0, ws.cell(row=snValue[x]+1, column=6).value)
				e7.insert(0, ws.cell(row=snValue[x]+1, column=7).value)
				ckvar8[a-1].set(ws.cell(row=snValue[x]+1, column=8).value)
				ckvar9[a-1].set(ws.cell(row=snValue[x]+1, column=9).value)
				ckvar10[a-1].set(ws.cell(row=snValue[x]+1, column=10).value)
				ckvar11[a-1].set(ws.cell(row=snValue[x]+1, column=11).value)
				tkvar3.set(ws.cell(row=snValue[x]+1, column=12).value)
				e13.insert(0, ws.cell(row=snValue[x]+1, column=13).value)
				
				if 1:
					e1[a-1].config(state='disabled', disabledforeground='#003366')
					e2.config(state='disabled', disabledforeground='#003366')
					e3[a-1].config(state='disabled', disabledforeground='#003366')
					e4.config(state='disabled', disabledforeground='#003366')
					e5.config(state='disabled', disabledforeground='#003366')
					e6.config(state='disabled', disabledforeground='#003366')
					e7.config(state='disabled', disabledforeground='#003366')
					c8.config(state='disabled', disabledforeground='#003366')
					c9.config(state='disabled', disabledforeground='#003366')
					c10.config(state='disabled', disabledforeground='#003366')
					c11.config(state='disabled', disabledforeground='#003366')
					e12.config(state='disabled', disabledforeground='#003366')
					e13.config(state='disabled', disabledforeground='#003366')
				
			del snValue[:]
		else:
			noResult = Label(newWindow, text="Nothing Found!")
			noResult.place(relx=0.36, rely=0.45, width=200, height=50)
			
		
	def saveData(x):
		ws.cell(row=lastRow, column=1, value=lastRow-1)
		ws.cell(row=lastRow, column=2, value=e2.get())
		ws.cell(row=lastRow, column=3, value=tkvar[a-1].get())
		ws.cell(row=lastRow, column=4, value=e4.get())
		ws.cell(row=lastRow, column=5, value=tkvar2.get())
		ws.cell(row=lastRow, column=6, value=e6.get())
		ws.cell(row=lastRow, column=7, value=e7.get())
		ws.cell(row=lastRow, column=8, value=ckvar8[a-1].get())
		ws.cell(row=lastRow, column=9, value=ckvar9[a-1].get())
		ws.cell(row=lastRow, column=10, value=ckvar10[a-1].get())
		ws.cell(row=lastRow, column=11, value=ckvar11[a-1].get())
		ws.cell(row=lastRow, column=12, value=tkvar3.get())
		if e13.get() == "":
			e13.insert(12, u"\u268a")
		ws.cell(row=lastRow, column=13, value=e13.get())
		
		wb.save('reg.xlsx')

		if x == 1:
			newEntryBtn.grid_forget()
			newEntry()
	
	welcomeText = "User: " + loginUser
	welcomeLabel = Label(newWindow, text=welcomeText, bg="#cfd8dc")
	if opt == 2:
		welcomeLabel = Label(canvas, text=welcomeText, bg="#cfd8dc")
	welcomeLabel.place(relx=0.01, rely=0.93)
	
	if opt == 1:
		saveDataBtn = Button(newWindow, text="Save Data", command=lambda: saveData(2))
		saveDataBtn.place(relx=0.83, rely=0.93)
			
		def callFunction():
			newWindow.destroy()
			openDashboard(newWindow, loginUser, loginPass)
		logOut = Button(newWindow, text="Go Back [Dashboard]", command=callFunction)
		logOut.place(relx=0.89, rely=0.93)
	
def openDashboard(window, loginUser, loginPass):
	# if loginUser == "nanc" and loginPass == "register":
	if loginUser in userGroup and loginPass == userDictionary[loginUser]:
			window.destroy()
			dashWindow = Frame(mainWindow, bg="#cfd8dc")
			dashWindow.place(anchor=NW, width=1200, height=800)
			bgImg = Image.open("dashBG.jpg")
			bgImage = Label(dashWindow)
			bgImage.img = ImageTk.PhotoImage(bgImg)
			bgImage.config(width = 1200, height = 530, image = bgImage.img)
			bgImage.place(anchor=NW)
			entryButton = Button(dashWindow, text="Enter Data", command=lambda: dataEntry(dashWindow,1))
			entryButton.place(relx=0.31, rely=0.70, width=200, height=50)
			
			def searchData():
				searchThis = searchEntry.get()
				# showinfo("Search Result","Under Development!")
				searchWindow = Toplevel(bg="#cfd8dc")
				searchWindow.title("Search Result")
				searchWindow.wm_iconbitmap('logo.ico')
				#pic = Image.open("searchBG.jpg")

				# searchBGImage = Label(searchWindow)
				# searchBGImage.img = ImageTk.PhotoImage(pic)
				# searchBGImage.config(image = searchBGImage.img)
				# searchBGImage.grid(row=0, column=0, rowspan=60, columnspan=13)
				
				# displayThis = 'You have searched for "' + searchThis + '"'
				# displayLabel = Label(searchWindow, text=displayThis, padx=10, pady=10)
				# displayLabel.place(relx=0.36, rely=0.05)
				
				dataEntry(searchWindow,2)
				 			
			global searchEntry
			searchEntry = Entry(dashWindow, bg="#f0f0f0", justify=CENTER)
			searchEntry.place(relx=0.31, rely=0.78, width=222, height=50)
			searchButton = Button(dashWindow, text="Search Data", command=lambda: searchData())
			searchButton.place(relx=0.50, rely=0.78, width=200, height=50)
			opnAB = Button(dashWindow, text="Open Address Book", command=lambda: os.startfile(r'\\NANCSERV\home\FreeAddressBook\AddressBook.exe'))
			opnAB.place(relx=0.50, rely=0.70, width=200, height=50)
			
			searchImg = Image.open("SearchImg.png")
			searchImage = Label(dashWindow)
			searchImage.img = ImageTk.PhotoImage(searchImg)
			searchImage.config(width = 40, height = 40, image = searchImage.img)
			searchImage.place(relx=0.315, rely=0.785)
			
			ck8 = Checkbutton(dashWindow, text="Recorded", bg="#cfd8dc")
			ck8.place(relx=0.31, rely=0.85)
			ck8.select()
			ck9 = Checkbutton(dashWindow, text="Checked", bg="#cfd8dc")
			ck9.place(relx=0.36, rely=0.85)
			ck9.select()
			ck10 = Checkbutton(dashWindow, text="Pending", bg="#cfd8dc")
			ck10.place(relx=0.42, rely=0.85)
			ck10.select()
			ck11 = Checkbutton(dashWindow, text="Completed", bg="#cfd8dc")
			ck11.place(relx=0.48, rely=0.85)
			ck11.select()
			
			welcomeText = "User: " + loginUser
			welcomeLabel = Label(dashWindow, text=welcomeText, bg="#cfd8dc")
			welcomeLabel.place(relx=0.01, rely=0.93)
			def callFunction():
				dashWindow.destroy()
				initiation()
			logOut = Button(dashWindow, text="Log Out", command=callFunction)
			logOut.place(relx=0.95, rely=0.93)
			
	else:
		showerror("Login Error","Wrong User Name or Password")

def initiation():
	#Login frame creation
	window = Frame(mainWindow)
	window.place(anchor=NW, width=1200, height=800)

	#Home image
	homeImg = Image.open("register_bg.jpg")
	homeImage = Label(window)
	homeImage.img = ImageTk.PhotoImage(homeImg)
	homeImage.config(width = 1200, height = 800, image = homeImage.img)
	homeImage.place(anchor=NW)

	#Login
	userName = Label(window, text="User Name", bg="#1E1F24", fg="#fff")
	userName.place(relx=0.38, rely=0.65)
	uName = Entry(window, bg="#1E1F24", fg="#fff")
	uName.place(relx=0.44, rely=0.65)
		
	passWord = Label(window, text="Password", bg="#1E1F24", fg="#fff")
	passWord.place(relx=0.38, rely=0.69)
	pWord = Entry(window, show="*", bg="#1E1F24", fg="#fff")
	pWord.place(relx=0.44, rely=0.69)
		
	def btnFunction(openOption):
		global loginUser, loginPass
		if openOption == 2:
			loginUser = "nanc"
			loginPass = "register"
		else:
			loginUser = uName.get()
			loginPass = pWord.get()
		openDashboard(window, loginUser, loginPass)

	#Dashboard Button
	dashboard = Button(window, text="Login", command=lambda: btnFunction(1), relief=RAISED, bg="#1E1F24", fg="#fff", activebackground="#ccc", activeforeground="#000")
	dashboard.place(relx=0.44, rely=0.73)
	
	#Dashboard Shortcut Button
	quickDashboard = Button(window, text="QLogin", command=lambda: btnFunction(2), relief=FLAT, bg="#fff", fg="#fff", activebackground="#ccc", activeforeground="#000")
	quickDashboard.place(relx=0.95, rely=0.93)

#Main window creation
mainWindow = Tk()
mainWindow.title("Register [NANC]")
mainWindow.geometry("1200x800")
mainWindow.wm_iconbitmap('logo.ico')

#Menubar creation
menubar = Menu(mainWindow)
filemenu = Menu(menubar, tearoff = 0)
filemenu.add_command(label = "Option One")
filemenu.add_command(label = "Option Two")
filemenu.add_separator()
filemenu.add_command(label = "Exit", command = mainWindow.destroy)
menubar.add_cascade(label = "File", menu = filemenu)

mainWindow.config(menu = menubar)
initiation()

mainWindow.mainloop()